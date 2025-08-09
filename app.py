"""Advanced PDF to Excel Data Converter - Converts PDF content into structured Excel data."""

import os
import tempfile
import io
import re
from datetime import datetime
from typing import List, Dict, Any, Tuple

from flask import (
    Flask,
    render_template,
    request,
    send_from_directory,
    redirect,
    url_for,
    flash,
)
from werkzeug.utils import secure_filename
from pdf2image import convert_from_path
from PIL import Image
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
import openpyxl.styles
import pandas as pd

# Advanced PDF processing libraries
import tabula
import camelot
import pytesseract

# ✅ Set Tesseract path
pytesseract.pytesseract.tesseract_cmd = r"C:\Program Files\Tesseract-OCR\tesseract.exe"

# ✅ Set Poppler path (PDF to image converter backend)
POPPLER_PATH = r"C:\poppler\poppler-24.08.0\Library\bin"

# 🔧 Folder configuration
UPLOAD_FOLDER = 'uploads'
OUTPUT_FOLDER = 'output'

# 🎯 File type restriction
ALLOWED_EXTENSIONS = {'pdf'}

# 🚀 Initialize the Flask app
app = Flask(__name__)
app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER
app.config['OUTPUT_FOLDER'] = OUTPUT_FOLDER
app.secret_key = 'supersecretkey'

# 📁 Ensure folders exist
os.makedirs(UPLOAD_FOLDER, exist_ok=True)
os.makedirs(OUTPUT_FOLDER, exist_ok=True)


def allowed_file(filename):
    """Check if the uploaded file has an allowed extension."""
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS


def extract_tables_with_tabula(pdf_path: str) -> List[pd.DataFrame]:
    """Extract tables using Tabula-py (works best with native PDF tables)."""
    try:
        print("[INFO] Extracting tables with Tabula...")
        tables = tabula.read_pdf(pdf_path, pages='all', multiple_tables=True)
        print(f"[INFO] Tabula found {len(tables)} tables")
        return tables
    except Exception as e:
        print(f"[WARNING] Tabula extraction failed: {e}")
        return []


def extract_tables_with_camelot(pdf_path: str) -> List[pd.DataFrame]:
    """Extract tables using Camelot-py (works best with bordered tables)."""
    try:
        print("[INFO] Extracting tables with Camelot...")
        tables = camelot.read_pdf(pdf_path, pages='all', flavor='stream')
        print(f"[INFO] Camelot found {len(tables)} tables")
        return [table.df for table in tables]
    except Exception as e:
        print(f"[WARNING] Camelot extraction failed: {e}")
        return []


def extract_text_with_ocr(image: Image.Image) -> str:
    """Extract text from image using OCR."""
    try:
        # OCR with optimized config
        config = r'--oem 3 --psm 6'
        text = pytesseract.image_to_string(image, config=config)
        return text.strip()
    except Exception as e:
        print(f"[WARNING] OCR extraction failed: {e}")
        return ""


def parse_table_from_text(text: str) -> List[List[str]]:
    """Parse table structure from OCR text."""
    try:
        lines = text.strip().split('\n')
        table_data = []
        
        for line in lines:
            if line.strip():
                # Split by common table separators
                cells = re.split(r'\s{2,}|\t|\|', line.strip())
                cells = [cell.strip() for cell in cells if cell.strip()]
                if cells:
                    table_data.append(cells)
        
        return table_data
    except Exception as e:
        print(f"[WARNING] Table parsing failed: {e}")
        return []


def detect_table_patterns_in_text(text: str) -> List[List[str]]:
    """Detect table patterns in text using regex and spacing analysis."""
    try:
        lines = text.strip().split('\n')
        potential_tables = []
        current_table = []
        
        for line in lines:
            line = line.strip()
            if not line:
                continue
            
            # Check if line has table-like characteristics
            # Multiple columns separated by spaces/tabs
            if re.search(r'\s{2,}', line) or '\t' in line:
                # Split by multiple spaces or tabs
                cells = re.split(r'\s{2,}|\t', line)
                cells = [cell.strip() for cell in cells if cell.strip()]
                
                if len(cells) >= 2:  # At least 2 columns
                    current_table.append(cells)
                else:
                    # End of current table
                    if len(current_table) >= 2:  # At least 2 rows
                        potential_tables.append(current_table)
                    current_table = []
            else:
                # End of current table
                if len(current_table) >= 2:
                    potential_tables.append(current_table)
                current_table = []
        
        # Add the last table if it exists
        if len(current_table) >= 2:
            potential_tables.append(current_table)
        
        return potential_tables
    except Exception as e:
        print(f"[WARNING] Table pattern detection failed: {e}")
        return []


def extract_structured_data_from_text(text: str) -> Dict[str, Any]:
    """Extract structured data from text using pattern matching."""
    structured_data = {
        'headers': [],
        'lists': [],
        'paragraphs': [],
        'financial_data': []
    }
    
    lines = text.strip().split('\n')
    
    for line in lines:
        line = line.strip()
        if not line:
            continue
            
        # Detect headers (short lines, often in caps)
        if len(line) < 100 and (line.isupper() or line.startswith(('Chapter', 'Section', 'Part'))):
            structured_data['headers'].append(line)
        
        # Detect lists (bullet points, numbers)
        elif re.match(r'^[\•\-\*\d]+\.?\s', line):
            structured_data['lists'].append(line)
        
        # Detect financial data (amounts, dates, etc.)
        elif re.search(r'\$[\d,]+\.?\d*|\d{1,2}/\d{1,2}/\d{2,4}', line):
            structured_data['financial_data'].append(line)
        
        # Long lines are paragraphs
        elif len(line) > 50:
            structured_data['paragraphs'].append(line)
    
    return structured_data


def create_consolidated_excel(pdf_path: str, all_data: Dict[str, Any]) -> Workbook:
    """Create a consolidated Excel workbook with better organization."""
    wb = Workbook()
    wb.remove(wb.active)  # Remove default sheet
    
    # Create main data sheet
    ws_main = wb.create_sheet("Data")
    
    current_row = 1
    
    # Process all tables and consolidate them
    all_tables = []
    
    # Add Tabula tables
    for i, table in enumerate(all_data.get('tabula_tables', [])):
        if not table.empty:
            all_tables.append(("Tabula Table", table))
    
    # Add Camelot tables
    for i, table in enumerate(all_data.get('camelot_tables', [])):
        if not table.empty:
            all_tables.append(("Camelot Table", table))
    
    # Add OCR tables
    for i, table in enumerate(all_data.get('ocr_tables', [])):
        if table:
            all_tables.append(("OCR Table", pd.DataFrame(table)))
    
    # If we have tables, add them to the main sheet
    if all_tables:
        for table_name, table in all_tables:
            # Add table data directly without header
            for row_idx, row_data in enumerate(table.values, 1):
                # Skip completely empty rows
                if not all(pd.isna(cell) for cell in row_data):
                    for col_idx, cell_value in enumerate(row_data, 1):
                        ws_main.cell(row=current_row, column=col_idx, value=cell_value)
                    current_row += 1
    
    # Add headers if any (as a simple list)
    if all_data.get('headers'):
        for header in all_data['headers']:
            if header.strip():  # Only add non-empty headers
                ws_main.cell(row=current_row, column=1, value=header)
                current_row += 1
    
    # Add financial data if any (as a simple list)
    if all_data.get('financial_data'):
        for item in all_data['financial_data']:
            if item.strip():  # Only add non-empty items
                ws_main.cell(row=current_row, column=1, value=item)
                current_row += 1
    
    # Add lists if any (as a simple list)
    if all_data.get('lists'):
        for item in all_data['lists']:
            if item.strip():  # Only add non-empty items
                ws_main.cell(row=current_row, column=1, value=item)
                current_row += 1
    
    # Add paragraphs if any (as a simple list)
    if all_data.get('paragraphs'):
        for paragraph in all_data['paragraphs']:
            if paragraph.strip():  # Only add non-empty paragraphs
                ws_main.cell(row=current_row, column=1, value=paragraph)
                current_row += 1
    
    # Create summary sheet
    ws_summary = wb.create_sheet("Summary")
    ws_summary.cell(row=1, column=1, value="PDF to Excel Data Conversion Summary")
    ws_summary.cell(row=1, column=1).font = openpyxl.styles.Font(bold=True, size=16)
    
    # Add summary statistics
    row = 3
    ws_summary.cell(row=row, column=1, value="Content Type")
    ws_summary.cell(row=row, column=2, value="Count")
    ws_summary.cell(row=row, column=1).font = openpyxl.styles.Font(bold=True)
    ws_summary.cell(row=row, column=2).font = openpyxl.styles.Font(bold=True)
    
    row += 1
    ws_summary.cell(row=row, column=1, value="Total Tables")
    ws_summary.cell(row=row, column=2, value=len(all_tables))
    
    row += 1
    ws_summary.cell(row=row, column=1, value="Headers")
    ws_summary.cell(row=row, column=2, value=len([h for h in all_data.get('headers', []) if h.strip()]))
    
    row += 1
    ws_summary.cell(row=row, column=1, value="Lists")
    ws_summary.cell(row=row, column=2, value=len([l for l in all_data.get('lists', []) if l.strip()]))
    
    row += 1
    ws_summary.cell(row=row, column=1, value="Financial Data")
    ws_summary.cell(row=row, column=2, value=len([f for f in all_data.get('financial_data', []) if f.strip()]))
    
    # Add conversion timestamp
    row += 2
    ws_summary.cell(row=row, column=1, value="Conversion Date")
    ws_summary.cell(row=row, column=2, value=datetime.now().strftime("%Y-%m-%d %H:%M:%S"))
    
    return wb


def process_pdf_to_structured_data(pdf_path: str) -> Dict[str, Any]:
    """Process PDF and extract structured data using multiple methods."""
    all_data = {
        'tabula_tables': [],
        'camelot_tables': [],
        'ocr_tables': [],
        'headers': [],
        'lists': [],
        'paragraphs': [],
        'financial_data': []
    }
    
    try:
        print("[INFO] Starting comprehensive PDF data extraction...")
        
        # Method 1: Tabula-py for native PDF tables
        tabula_tables = extract_tables_with_tabula(pdf_path)
        all_data['tabula_tables'] = tabula_tables
        
        # Method 2: Camelot-py for bordered tables
        camelot_tables = extract_tables_with_camelot(pdf_path)
        all_data['camelot_tables'] = camelot_tables
        
        # Method 3: OCR + Pattern Detection for image-based tables
        print("[INFO] Converting PDF to images for OCR analysis...")
        images = convert_from_path(
            pdf_path, 
            fmt="PNG",
            dpi=200,
            poppler_path=POPPLER_PATH
        )
        
        ocr_tables = []
        structured_text_data = {
            'headers': [],
            'lists': [],
            'paragraphs': [],
            'financial_data': []
        }
        
        for page_num, img in enumerate(images):
            print(f"[INFO] Processing page {page_num + 1} with OCR...")
            
            # Extract text from image
            full_text = extract_text_with_ocr(img)
            
            # Detect table patterns in the text
            page_tables = detect_table_patterns_in_text(full_text)
            ocr_tables.extend(page_tables)
            
            # Extract structured data from text
            page_structured_data = extract_structured_data_from_text(full_text)
            
            # Merge structured data
            for key in structured_text_data:
                structured_text_data[key].extend(page_structured_data[key])
        
        all_data['ocr_tables'] = ocr_tables
        all_data.update(structured_text_data)
        
        print(f"[INFO] Extraction complete:")
        print(f"  - Tabula tables: {len(tabula_tables)}")
        print(f"  - Camelot tables: {len(camelot_tables)}")
        print(f"  - OCR tables: {len(ocr_tables)}")
        print(f"  - Headers: {len(structured_text_data['headers'])}")
        print(f"  - Lists: {len(structured_text_data['lists'])}")
        print(f"  - Financial data: {len(structured_text_data['financial_data'])}")
        
        return all_data
        
    except Exception as e:
        print(f"[ERROR] PDF processing failed: {str(e)}")
        raise


@app.route('/', methods=['GET', 'POST'])
def upload_file():
    """Handle file upload and trigger PDF processing."""
    if request.method == 'POST':
        if 'file' not in request.files:
            flash('No file part')
            return redirect(request.url)
        file = request.files['file']
        if file.filename == '':
            flash('No selected file')
            return redirect(request.url)
        if file and allowed_file(file.filename):
            filename = secure_filename(file.filename)
            filepath = os.path.join(app.config['UPLOAD_FOLDER'], filename)
            file.save(filepath)
            try:
                excel_filename = process_pdf(filepath)
                return redirect(url_for('download_file', filename=excel_filename))
            except Exception as e:
                print("[ERROR]", str(e))
                flash(f'Processing failed: {e}')
                return redirect(request.url)
    return render_template('index.html')


def process_pdf(path):
    """Main PDF processing function."""
    try:
        # Extract structured data from PDF
        structured_data = process_pdf_to_structured_data(path)
        
        # Create Excel workbook with consolidated data
        excel_name = os.path.splitext(os.path.basename(path))[0] + "_converted.xlsx"
        excel_path = os.path.join(app.config["OUTPUT_FOLDER"], excel_name)
        
        wb = create_consolidated_excel(path, structured_data)
        wb.save(excel_path)

        print(f"[INFO] Excel generated: {excel_path}")
        return excel_name
        
    except Exception as e:
        print(f"[ERROR] PDF processing failed: {str(e)}")
        raise


@app.route('/output/<filename>')
def download_file(filename):
    """Serve the generated Excel file for download."""
    return send_from_directory(app.config['OUTPUT_FOLDER'], filename, as_attachment=True)


if __name__ == '__main__':
    app.run(debug=True)
